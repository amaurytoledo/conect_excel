import openpyxl
import tkinter as tk

root = tk.Tk()
root.title("Adicionar dados dos jogadores na Planilha")

def adicionar_dados():
    wb = openpyxl.load_workbook('jogadores.xlsx')
    sheet = wb.active
    nome = entry_nome.get()
    idade = entry_idade.get()
    posicao = entry_posicao.get()
    preco  = entry_preco.get()
    linha = sheet.max_row + 1
    sheet.cell(row=linha, column=1).value = nome
    sheet.cell(row=linha, column=2).value = idade
    sheet.cell(row=linha, column=3).value = posicao
    sheet.cell(row=linha, column=4).value = preco
    wb.save('jogadores.xlsx')
    entry_nome.delete(0, tk.END)
    entry_idade.delete(0, tk.END)
    entry_posicao.delete(0, tk.END)
    entry_preco.delete(0, tk.END)

label_nome = tk.Label(root, text="Nome:")
label_nome.grid(row=0, column=0, padx=10, pady=10)
entry_nome = tk.Entry(root, width=20)
entry_nome.grid(row=0, column=1, padx=10, pady=10)

label_idade = tk.Label(root, text="Idade:")
label_idade.grid(row=1, column=0, padx=10, pady=10)
entry_idade = tk.Entry(root, width=20)
entry_idade.grid(row=1, column=1, padx=10, pady=10)

label_posicao = tk.Label(root, text="Posição:")
label_posicao.grid(row=2, column=0, padx=10, pady=10)
entry_posicao = tk.Entry(root, width=20)
entry_posicao.grid(row=2, column=1, padx=10, pady=10)

label_preco = tk.Label(root, text="Preço:")
label_preco.grid(row=3, column=0, padx=10, pady=10)
entry_preco = tk.Entry(root, width=20)
entry_preco.grid(row=3, column=1, padx=10, pady=10)

button_adicionar = tk.Button(root, text="Adicionar", command=adicionar_dados)
button_adicionar.grid(row=4, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()
